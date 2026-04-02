#!/usr/bin/env python3
"""
多框架自动化压测工具（统一版）
支持 FastDeploy / SGLang 两套框架配置

用法:
    python run_bench.py                          # 交互菜单，使用当前目录下 bench.yaml
    python run_bench.py --config fd_bench.yaml   # 指定 FD 配置
    python run_bench.py --config sglang_bench.yaml  # 指定 SGLang 配置
    python run_bench.py --kill                   # 只 kill 当前服务
    python run_bench.py --smoke-test             # 冒烟测试：--num-prompts 替换为 10
"""

import glob
import os
import re
import signal
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import yaml

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[error] 请先安装: pip install openpyxl")

# ── 指标展示配置 ──────────────────────────────────────────────────────────────
PRIORITY_COLS = [
    ("Mean Input Length",               "平均输入长度 (tok)"),
    ("Mean Output Length",              "平均输出长度 (tok)"),
    ("Request throughput (req/s)",      "QPS (req/s)"),
    ("Output token throughput (tok/s)", "TPS (tok/s)"),
    ("Mean Decode",                     "平均解码速度 (tok/s)"),
    ("Mean TTFT (ms)",                  "首token均值时延 (ms)"),
    ("Mean E2EL (ms)",                  "整句均值时延 (ms)"),
]
PRIORITY_KEYS  = [k for k, _ in PRIORITY_COLS]
PRIORITY_LABEL = {k: v for k, v in PRIORITY_COLS}

MEAN_ONLY_GROUPS = {"Input Length", "Output Length", "Cached Tokens"}
SKIP_GROUPS      = {"S_TTFT", "S_ITL", "S_E2EL"}

EXTRA_ORDER = [
    "Successful requests", "Benchmark duration (s)",
    "Total input tokens", "Total generated tokens",
    "Total Token throughput (tok/s)",
    "Median Decode", "P80 Decode", "P95 Decode", "P99 Decode", "P99.9 Decode",
    "Median TTFT (ms)", "P80 TTFT (ms)", "P95 TTFT (ms)",
    "P99 TTFT (ms)", "P99.9 TTFT (ms)", "P99.95 TTFT (ms)", "P99.99 TTFT (ms)",
    "Mean TPOT (ms)", "P80 TPOT (ms)", "P95 TPOT (ms)", "P99 TPOT (ms)", "P99.9 TPOT (ms)",
    "Mean ITL (ms)", "P80 ITL (ms)", "P95 ITL (ms)", "P99 ITL (ms)", "P99.9 ITL (ms)",
    "Median E2EL (ms)", "P80 E2EL (ms)", "P95 E2EL (ms)",
    "P99 E2EL (ms)", "P99.9 E2EL (ms)", "P99.95 E2EL (ms)", "P99.99 E2EL (ms)",
]

# 各框架 ready 标志（可在 YAML global.ready_marker 覆盖）
READY_MARKERS = {
    "fd":     "Application startup complete",
    "sglang": "Application startup complete",
}


def get_framework_version(framework: str, experiments: list) -> dict:
    """从实验配置的 server 命令中自动提取框架源码路径，获取版本和 git commit"""
    info = {"version": "unknown", "commit": "unknown", "commit_short": "unknown", "extra": ""}

    first_server_cmd = ""
    for exp in experiments:
        cmd = exp.get("server", "").strip()
        if cmd:
            first_server_cmd = cmd
            break
    if not first_server_cmd:
        return info

    if framework == "fd":
        repo_path = _extract_path(first_server_cmd, r'PYTHONPATH="?([^":\s]+FastDeploy)')
        if repo_path:
            info.update(_git_info(repo_path))
            vf = Path(repo_path) / "fastdeploy" / "version.txt"
            if vf.exists():
                info["extra"] = vf.read_text().strip()

    elif framework == "sglang":
        repo_path = _extract_path(first_server_cmd, r'PYTHONPATH=([^":\s]+sglang/python)')
        if repo_path:
            repo_path = str(Path(repo_path).parent)
        if repo_path:
            info.update(_git_info(repo_path))
        python_bin = _extract_path(first_server_cmd, r'(/\S+/bin/python\S*)')
        if python_bin:
            ver = _run_quiet([python_bin, "-c",
                              "from sglang.version import __version__; print(__version__)"])
            if ver:
                info["version"] = ver

    return info


def _extract_path(cmd: str, pattern: str) -> str | None:
    for line in cmd.splitlines():
        m = re.search(pattern, line)
        if m and Path(m.group(1)).exists():
            return m.group(1)
    return None


def _git_info(repo_path: str) -> dict:
    result = {}
    commit = _run_quiet(["git", "-C", repo_path, "rev-parse", "HEAD"])
    if commit:
        result["commit"] = commit
        result["commit_short"] = commit[:9]
    desc = _run_quiet(["git", "-C", repo_path, "describe", "--tags", "--always"])
    if desc:
        result["version"] = desc
    branch = _run_quiet(["git", "-C", repo_path, "rev-parse", "--abbrev-ref", "HEAD"])
    if branch:
        result["branch"] = branch
    return result


def _run_quiet(cmd: list, timeout: int = 10) -> str | None:
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return r.stdout.strip() if r.returncode == 0 else None
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return None


def should_include(key: str) -> bool:
    for grp in SKIP_GROUPS:
        if grp in key:
            return False
    for grp in MEAN_ONLY_GROUPS:
        if grp in key and not key.startswith("Mean"):
            return False
    return True


# ── 解析压测输出 ──────────────────────────────────────────────────────────────
def parse_output(text: str) -> dict:
    metrics = {}
    in_summary = False
    for line in text.splitlines():
        stripped = line.strip()
        if "Serving Benchmark Result" in stripped:
            in_summary = True
            continue
        if not in_summary:
            continue
        if re.match(r"^=+$", stripped):
            break
        m = re.match(r"^(.+?):\s+([\d.]+)\s*$", stripped)
        if m:
            try:
                metrics[m.group(1).strip()] = float(m.group(2))
            except ValueError:
                pass
    return metrics


# ── 等待服务就绪 ──────────────────────────────────────────────────────────────
def wait_for_server(log_file: Path, timeout: int, ready_marker: str) -> bool:
    print(f"  等待服务就绪（监听 {log_file.name}），超时 {timeout}s ...", flush=True)
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            text = log_file.read_text(errors="replace")
            if ready_marker in text:
                print("  ✓ 服务已就绪", flush=True)
                return True
        except FileNotFoundError:
            pass
        remaining = int(deadline - time.time())
        print(f"  ... 等待启动，剩余 {remaining}s", end="\r", flush=True)
        time.sleep(3)
    print("\n  ✗ 等待超时", flush=True)
    return False


# ── 启动服务（后台） ──────────────────────────────────────────────────────────
def start_server(cmd: str, log_file: Path) -> subprocess.Popen:
    with open(log_file, "w") as f:
        proc = subprocess.Popen(
            ["bash", "-c", cmd],
            stdout=f,
            stderr=f,
            preexec_fn=os.setsid,
        )
    return proc


# ── Kill 服务 ─────────────────────────────────────────────────────────────────
def kill_server(port: int, proc: subprocess.Popen = None,
                extra_ports: list = None, cuda_devices: str = None,
                framework: str = "fd"):
    all_ports = [port] + [port + offset for offset in (extra_ports or [])]
    print(f"  kill 服务（端口 {all_ports}）...", flush=True)
    killed = set()

    # 1. 按端口 kill
    for p in all_ports:
        r = subprocess.run(f"lsof -ti :{p}", shell=True, capture_output=True, text=True)
        for pid in r.stdout.strip().splitlines():
            pid = pid.strip()
            if pid:
                subprocess.run(f"kill -9 {pid}", shell=True)
                killed.add(pid)

    # 2. 按指定 GPU 上的占用进程 kill（仅 FD 需要，sglang 不传 cuda_devices 即跳过）
    if cuda_devices:
        gpu_ids = [x.strip() for x in str(cuda_devices).split(",") if x.strip()]
        for gid in gpu_ids:
            r = subprocess.run(
                f"nvidia-smi --query-compute-apps=pid --format=csv,noheader --id={gid}",
                shell=True, capture_output=True, text=True,
            )
            for pid in r.stdout.strip().splitlines():
                pid = pid.strip()
                if pid and pid not in killed:
                    subprocess.run(f"kill -9 {pid}", shell=True)
                    killed.add(pid)

    # 3. 兜底：kill 脚本进程组
    if proc and proc.poll() is None:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except (ProcessLookupError, OSError):
            pass

    if killed:
        print(f"  已 kill PID: {', '.join(sorted(killed))}", flush=True)
    else:
        print("  未发现需要 kill 的进程", flush=True)

    # 4. 清理 FD Unix domain socket 文件（仅 FD 框架需要）
    if framework == "fd":
        sock_candidates = [f"/dev/shm/fd_task_queue_{p}.sock" for p in all_ports]
        sock_candidates += glob.glob("/dev/shm/fd_*.sock")
        cleaned = set()
        for sock in sock_candidates:
            if os.path.exists(sock):
                try:
                    os.remove(sock)
                    cleaned.add(sock)
                except OSError:
                    pass
        if cleaned:
            print(f"  已清理 socket 文件: {', '.join(sorted(cleaned))}", flush=True)


# ── 运行压测 ──────────────────────────────────────────────────────────────────
def run_infer(cmd: str, log_file: Path, run_dir: Path) -> str:
    print(f"  运行压测，输出实时可见 → {log_file.name}", flush=True)
    with open(log_file, "w") as lf:
        proc = subprocess.Popen(
            ["bash", "-c", cmd],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            cwd=str(run_dir),
        )
        lines = []
        for line in proc.stdout:
            sys.stdout.write(line)
            sys.stdout.flush()
            lf.write(line)
            lines.append(line)
        proc.wait()
    return "".join(lines)


# ── 写 Excel ─────────────────────────────────────────────────────────────────
def write_excel(all_results: list, out_path: Path):
    seen_keys = []
    for r in all_results:
        for k in r.get("metrics", {}):
            if k not in seen_keys and should_include(k):
                seen_keys.append(k)

    ordered_keys = []
    for k in PRIORITY_KEYS:
        if k in seen_keys and k not in ordered_keys:
            ordered_keys.append(k)
    for k in EXTRA_ORDER:
        if k in seen_keys and k not in ordered_keys:
            ordered_keys.append(k)
    for k in seen_keys:
        if k not in ordered_keys:
            ordered_keys.append(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    fixed_cols  = ["框架", "实验名称", "版本", "Commit", "运行时间", "状态", "起服务脚本", "起请求脚本"]
    metric_cols = [PRIORITY_LABEL.get(k, k) for k in ordered_keys]
    header = fixed_cols + metric_cols

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    fail_fill = PatternFill("solid", fgColor="FCE4D6")
    for ri, r in enumerate(all_results, 2):
        status   = r["status"]
        row_fill = ok_fill if status == "ok" else fail_fill

        ws.cell(row=ri, column=1, value=r.get("framework", ""))
        ws.cell(row=ri, column=2, value=r["name"])
        ws.cell(row=ri, column=3, value=r.get("version", ""))
        ws.cell(row=ri, column=4, value=r.get("commit", ""))
        ws.cell(row=ri, column=5, value=r["time"])
        ws.cell(row=ri, column=6, value=status)
        ws.cell(row=ri, column=7, value=r.get("server_cmd", ""))
        ws.cell(row=ri, column=8, value=r.get("infer_cmd", ""))

        for ci, k in enumerate(ordered_keys, len(fixed_cols) + 1):
            ws.cell(row=ri, column=ci, value=r.get("metrics", {}).get(k))

        for ci in range(1, len(header) + 1):
            ws.cell(row=ri, column=ci).fill = row_fill

    for ci, h in enumerate(header, 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(h)),
            *(len(str(ws.cell(row=ri, column=ci).value or ""))
              for ri in range(2, len(all_results) + 2)),
        )
        limit = 60 if ci in (7, 8) else 30
        ws.column_dimensions[col_letter].width = min(max_len + 2, limit)

    ws.freeze_panes = "I2"
    wb.save(out_path)
    print(f"\n[Excel] 已保存: {out_path}")


# ── 交互菜单 ──────────────────────────────────────────────────────────────────
def show_menu(experiments: list, framework: str) -> list:
    print("\n" + "=" * 60)
    print(f"  自动化压测  [{framework}]")
    print("=" * 60)
    print(f"  [0] 全部运行 ({len(experiments)} 个实验)")
    for i, exp in enumerate(experiments, 1):
        print(f"  [{i:2d}] {exp['name']}")
    print("  ─────────────────────────────────────────────────────")
    print("  [k] Kill 当前服务")
    print("  [q] 退出")
    print("=" * 60)
    raw = input("选择（多个用逗号，如 1,3）: ").strip().lower()

    if raw in ("q", "quit"):
        sys.exit(0)
    if raw == "k":
        return "kill"
    if raw in ("0", "all"):
        return list(range(len(experiments)))
    indices = []
    for part in raw.split(","):
        part = part.strip()
        if part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < len(experiments):
                indices.append(idx)
            else:
                print(f"  [warn] 序号 {part} 超出范围，忽略")
        else:
            print(f"  [warn] 无效输入 '{part}'，忽略")
    return indices


# ── 主流程 ────────────────────────────────────────────────────────────────────
def main():
    # 清理代理环境变量，避免 localhost 请求被 Squid 等代理拦截
    for key in ("http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY"):
        os.environ.pop(key, None)

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=Path(__file__).parent / "bench.yaml",
                        help="YAML 配置文件路径（默认 bench.yaml）")
    parser.add_argument("--kill", action="store_true", help="只 kill 当前服务")
    parser.add_argument("--smoke-test", action="store_true",
                        help="冒烟测试：将所有 --num-prompts 替换为 10")
    args = parser.parse_args()

    cfg_path = Path(args.config)
    if not cfg_path.exists():
        sys.exit(f"[error] 配置文件不存在: {cfg_path}")

    with open(cfg_path) as f:
        cfg = yaml.safe_load(f)

    g = cfg.get("global", {})

    # 从配置文件名自动推断框架名，也可在 global.framework 里显式指定
    framework = g.get("framework") or (
        "fd" if "fd" in cfg_path.stem.lower() else
        "sglang" if "sglang" in cfg_path.stem.lower() else
        cfg_path.stem
    )

    port              = g.get("port", 2786)
    extra_ports       = g.get("extra_ports_offsets", [])
    ready_timeout     = g.get("server_ready_timeout", 300)
    shutdown_wait     = g.get("shutdown_wait", 20)
    cuda_devices      = g.get("CUDA_VISIBLE_DEVICES", None)
    results_dir       = Path(cfg_path.parent) / g.get("results_dir", "./results")
    ready_marker      = g.get("ready_marker", READY_MARKERS.get(framework,
                              "Application startup complete"))

    experiments = cfg.get("experiments", [])

    if args.kill:
        kill_server(port, extra_ports=extra_ports, cuda_devices=cuda_devices, framework=framework)
        return

    selection = show_menu(experiments, framework)
    if selection == "kill":
        kill_server(port, extra_ports=extra_ports, cuda_devices=cuda_devices, framework=framework)
        return
    if not selection:
        print("未选择任何实验，退出。")
        return

    # 获取框架版本信息
    fw_ver = get_framework_version(framework, experiments)
    ver_line = f"{fw_ver['version']}  commit={fw_ver['commit_short']}"
    if fw_ver.get("branch"):
        ver_line += f"  branch={fw_ver['branch']}"
    print(f"\n[版本] {framework}: {ver_line}")
    if fw_ver.get("extra"):
        for line in fw_ver["extra"].splitlines():
            print(f"        {line}")

    run_id  = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = results_dir / f"{framework}_{run_id}"
    run_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n框架: {framework}  结果目录: {run_dir}\n")

    # 保存版本信息到结果目录
    with open(run_dir / "version_info.txt", "w") as f:
        f.write(f"framework: {framework}\n")
        f.write(f"version: {fw_ver['version']}\n")
        f.write(f"commit: {fw_ver['commit']}\n")
        if fw_ver.get("branch"):
            f.write(f"branch: {fw_ver['branch']}\n")
        if fw_ver.get("extra"):
            f.write(f"\n{fw_ver['extra']}\n")

    all_results = []
    excel_path  = run_dir / f"{framework}_bench_{run_id}.xlsx"

    for idx in selection:
        exp  = experiments[idx]
        name = exp["name"]
        print(f"\n{'─'*60}")
        print(f"[{idx+1}/{len(experiments)}] [{framework}] 实验: {name}")
        print(f"{'─'*60}")

        server_log = run_dir / f"{idx+1}_{name}_server.log"
        infer_log  = run_dir / f"{idx+1}_{name}_infer.log"

        # 1. 启动服务
        server_cmd = exp.get("server", "").strip()
        proc = None
        if server_cmd:
            if args.smoke_test:
                server_cmd_display = server_cmd
            # FD 框架：为每个实验单独保存日志，避免被后续实验覆盖
            if framework == "fd":
                fd_log_dir = run_dir / f"{idx+1}_{name}_fd_log"
                fd_log_dir.mkdir(parents=True, exist_ok=True)
                server_cmd = f"export FD_LOG_DIR={fd_log_dir}\n" + server_cmd
                print(f"  FD 日志目录 → {fd_log_dir.name}", flush=True)
            print(f"  预清理残留进程...", flush=True)
            kill_server(port, extra_ports=extra_ports, cuda_devices=cuda_devices, framework=framework)
            time.sleep(3)
            print(f"  启动服务 → {server_log.name}", flush=True)
            proc = start_server(server_cmd, server_log)
            ready = wait_for_server(server_log, ready_timeout, ready_marker)
            if not ready:
                print(f"  [SKIP] 服务启动超时，跳过实验 {name}")
                kill_server(port, proc, extra_ports=extra_ports, cuda_devices=cuda_devices, framework=framework)
                all_results.append({
                    "framework": framework,
                    "name": name,
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "server_timeout",
                    "server_cmd": server_cmd,
                    "infer_cmd": "",
                    "metrics": {},
                })
                write_excel(all_results, excel_path)
                print(f"  已实时写入 Excel: {excel_path.name}", flush=True)
                continue
        else:
            print("  server 为空，跳过启动（假设服务已在运行）")

        # 2. 运行压测
        infer_cmd = exp.get("infer", "").strip()
        if infer_cmd and args.smoke_test:
            infer_cmd = re.sub(r"(--num-prompts\s+)\d+", r"\g<1>10", infer_cmd)
            print("  [smoke-test] --num-prompts 已替换为 10", flush=True)
        output = ""
        if infer_cmd:
            output = run_infer(infer_cmd, infer_log, run_dir)
        else:
            print("  infer 为空，跳过压测")

        # 3. 解析结果
        metrics = parse_output(output)
        status  = "ok" if metrics.get("Successful requests", 0) > 0 else "failed"
        if metrics:
            print(f"  ✓ {len(metrics)} 项指标  "
                  f"QPS={metrics.get('Request throughput (req/s)', 'N/A')}  "
                  f"TTFT={metrics.get('Mean TTFT (ms)', 'N/A')}ms")
        else:
            print("  ✗ 未解析到指标（压测可能失败）")

        all_results.append({
            "framework": framework,
            "name": name,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": status,
            "version": fw_ver.get("version", ""),
            "commit": fw_ver.get("commit_short", ""),
            "server_cmd": server_cmd,
            "infer_cmd": infer_cmd,
            "metrics": metrics,
        })

        # 每个实验后立即写 Excel
        write_excel(all_results, excel_path)
        print(f"  已实时写入 Excel: {excel_path.name}", flush=True)

        # 4. Kill 服务
        if server_cmd:
            kill_server(port, proc, extra_ports=extra_ports, cuda_devices=cuda_devices, framework=framework)
            print(f"  等待 GPU 显存释放 {shutdown_wait}s ...", flush=True)
            time.sleep(shutdown_wait)

    print(f"\n{'='*60}")
    print(f"  全部完成  [{framework}]  共 {len(all_results)} 个实验")
    print(f"  Excel: {excel_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()

# 后台运行示例:
# echo "0" | nohup python run_bench.py --config fd_bench.yaml > out_fd.txt 2>&1 &
# echo "0" | nohup python run_bench.py --config sglang_bench.yaml > out_sglang.txt 2>&1 &
